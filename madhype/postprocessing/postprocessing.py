
"""
Attempt to test variable cells per well
"""

# standard libraries
from math import log10,ceil,floor
import copy
import random
import os

# nonstandard libraries
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl

# homegrown libraries
import plots
from ..defaults import general_options as default_options
plt.rcParams["font.family"] = "serif"

'''
MAIN FUNCTIONS
'''

#------------------------------------------------------------------------------# 

def analyze_results(results,data,**kwargs):

    """
    Processes results dictionary using data object as reference
    """

    print 'Analyzing results..'

    # default options parameters
    options = default_options.copy()
    
    # update options
    # kwargs must contain 'reference' if the true cells are not included in the data
    options.update(kwargs)

    # if this is not a simulated dataset with a hard set of 
    if not 'cells' in data:

        if options['reference'] == None:
            print 'No cell or subject reference provided, cannot produce meaningful results!'
            return {'raw_results': results}

        if not is_reference_valid(options):
            print 'Reference not valid, cannot perform analysis!'
            return {'raw_results': results}
        
        print 'Starting analysis using subject reference'

        cresults = get_results_from_subject_reference(results,data,options)

    else:

        cresults = get_results_from_cell_reference(results,data,options)

    if options['visual']:

        print 'Generating figures...'

        ### AUROC FIGURE ###
        if options['plot_auroc']:
            plots.plot_auroc(cresults,**options)    

        ### FREQUENCY ESTIMATION FIGURE ###
        if options['plot_frequency_estimation']:
            plots.plot_frequency_estimation(cresults,**options)    

        ### REPERTOIRE DISPLAY FIGURE ###
        if options['plot_repertoire']:
            plots.plot_repertoire(cresults,**options)    
       
    return cresults


def get_results_from_subject_reference(results,data,options):

    reference = options['reference']

    # Randomize order of results so subsequent calculations are
    # not dependent on the order of the results list
    random.shuffle(results)

    cells_with_scores = sorted(results,key=lambda x: -x[1])
    cells_without_scores = [i[0] for i in cells_with_scores]

    reported_matches = [('TCRA','TCRB','p-ratio','TCRA origin','TCRB origin','f_ij','f_i','f_j')]
    total = 0
    x,y,n = 0,0,0
    x1,y1 = [0],[0]
    pattern = []

    for index,c in enumerate(cells_with_scores):

        if index < 1000:
            pass#print index,' : ',c 

        a,b = '',''

        i = c[0][0][0]
        j = c[0][1][0]

        if i in reference['A']['X']: a += 'X'
        if i in reference['A']['Y']: a += 'Y'
        if j in reference['B']['X']: b += 'X'
        if j in reference['B']['Y']: b += 'Y'

        try:
            reported_matches.append((i,j,c[1],a,b,c[2]['i'],c[2]['j'],c[2]['ij']))
        except KeyError:
            reported_matches.append((i,j,c[1],a,b))

        total += 1

        if (a == 'X' and b == 'X') or (a == 'Y' and b == 'Y'):
            pattern.append(1)
            y += 1
        elif (a == 'X' and b == 'Y') or (a == 'Y' and b == 'X'):
            pattern.append(-1)
            x1.append(x)
            y1.append(y)
            x += 1
            x1.append(x)
            y1.append(y)
        else:
            pattern.append(0)
            n += 1
            continue 

        if 2*x1[-1] > options['fdr']*y1[-1]:
            print 'FDR met!'
            break

    results = {
              'positives':y1[-1],
              'negatives':x1[-1],
              'neutrals':n,
              'pattern':pattern,
              'total':total,
              'x':x1,
              'y':y1,
              'options':options,
              'raw_results': cells_with_scores,
              'matches':reported_matches,
              }

    if options['save_to_excel']:
        write_results_to_xslx(results)

    return results

def write_results_to_xslx(original_results):

    # pull options and matches from results
    results = copy.copy(original_results)

    options = results.pop('options',{})
    matches = results.pop('matches',{})
    if matches == {}:
	 matches = results.pop('raw_results',{})
    options.pop('reference',{})
    results.pop('reference',{})

    # create new workbook
    wb = openpyxl.Workbook() # open new workbook

    # settings sheet
    _write_dict_to_wb(wb,options,'OPTIONS')
    _write_dict_to_wb(wb,results,'SUMMARY')

    # data sheet
    ws = wb.create_sheet(title='MATCHES')
    ws.append(('Match Labels (A,B)','Match Confidence Ratio (10^-x)','Clonal Frequencies (i->A,j->B)'))
    for match in matches:
        ws.append((str(m) for m in match))

    # save file to unique name
    _unique_wb_save(wb,'results','./results')

def _write_dict_to_wb(wb,settings,sheet_name):
    """ Settings writing """
    ws = wb.create_sheet(title=sheet_name)

    for k,v in settings.items():
	if k == 'raw_results': continue # skip a listing of matches
	if k == 'matches': continue # skip a listing of matches
        if isinstance(v,dict):
            ws.append((k,'dict->'))
            for k2,v2 in v.items():
                if isinstance(v2,(tuple,list)):
                    ws.append(('->',k2)+(str(v2),))
                else:
                    ws.append(('->',k2,v2))
        elif isinstance(v,(tuple,list)):
            ws.append([k] + [str(v)])
        else:
            ws.append((k,str(v)))

def _unique_wb_save(wb,name,directory):

    if not os.path.isdir(directory):
        os.mkdir(directory)

    # reference object
    for i in range(1,10000):

        reference_fname = os.path.join(directory,'{}_{}.xlsx'.format(name,str(i).zfill(4)))

        if os.path.isfile(reference_fname): continue

        # try removing starting blank sheet
        try: del wb['Sheet']
        except: pass

        print 'Saving to excel:',reference_fname

        wb.save(reference_fname) # save final workbook

        return

def get_results_from_cell_reference(results,data,options):

    random.shuffle(results)

    # these are guessed cells
    cells_with_scores = sorted(results,key=lambda x: -x[1])
    cells_without_scores = [i[0] for i in cells_with_scores]

    # these are actual cells
    cells_temp = sorted([((tuple(a),tuple(b)),f) for (a,b),f in data['pairs']],key=lambda x: -x[1])
    cells_record = set([c[0] for c in cells_temp])
    cells_label = [c[0] for c in cells_temp]
    cells_freqs = [c[1] for c in cells_temp]
    total_cells = len(cells_temp)

    # look at error rate (stratified by confidence)
    x1,y1 = [0],[0]
    dx,dy = 0,0
    warning=True # warning if you didn't pass enough points

    for i in range(len(cells_with_scores)):
        c, score, _ = cells_with_scores[i]

        if c in cells_record:
            dy += 1
        else:
            dx += 1

        if i == len(cells_with_scores)-1 or score != cells_with_scores[i+1][1]:
            if x1[-1]+dx > options['fdr'] * (y1[-1]+dy):
                # Determine where the ROC curve passes the FDR cutoff line
                t = - float(options['fdr']*y1[-1] - x1[-1]) / (options['fdr']*dy - dx)
                x1.append(x1[-1] + int(dx*t))
                y1.append(y1[-1] + int(dy*t))
                warning = False
                break
            else:
                x1.append(x1[-1]+dx)
                y1.append(y1[-1]+dy)
                dx,dy = 0,0
#           
#    for c in cells_with_scores:
#        try:
#            data['cells'][c[0]]
#            x1.append(x1[-1])
#            y1.append(y1[-1]+1)
#        except KeyError:
#            x1.append(x1[-1]+1)
#            y1.append(y1[-1])
#        if x1[-1] > options['fdr']*y1[-1]:
#            warning=False
#            break
    
    if warning:
        print 'WARNING: Number of passed guesses did not meet FDR!'
        print 'Lower the match threshold and you will see better results!'

    total_matches_at_fdr = x1[-1] + y1[-1]

    # NEW METHOD OF CALCULATING REPERTOIRE FRACTION, W/ COVERAGE FOR DUAL CLONES
    cells_without_scores_at_fdr = set(cells_without_scores[:total_matches_at_fdr])

    # create dict with complete cells and the number of matches required to "finish" identifying cell
    complete_cells = {(tuple(a),tuple(b)): f for (a,b),f in data['cells']}
    complete_cell_match_counter = {k: len(k[0])*len(k[1]) for k in complete_cells}
    map_matches_to_complete_cells = {}

    # create a mapping from match pairs to full clones
    for k in complete_cell_match_counter:
        for a in k[0]:
            for b in k[1]:
                try:
                    map_matches_to_complete_cells[((a,),(b,))].append(k)
                except KeyError:
                    map_matches_to_complete_cells[((a,),(b,))] = [k]

    # count the number of matches that fit each complete cell
    for guessed_cell_match in cells_without_scores_at_fdr:
        try:
            for complete_cell_match in map_matches_to_complete_cells[guessed_cell_match]:
                complete_cell_match_counter[complete_cell_match] += -1
        except KeyError:
            pass

    freqs = []

    for k,v in complete_cell_match_counter.items():
        if v == 0:
            freqs.append(complete_cells[k])
        if v < 0:
            print 'We have a problem... {}:{}'.format(k,v)

    frac_repertoire = sum(freqs)

    #print 'New frac repertoire:',frac_repertoire

    # This method wasn't compatible with dual clones (since you need multiple matches for one cell)
    '''
    freqs = [cells_freqs[i] for i,c in enumerate(cells_label) \
             if c in cells_without_scores_at_fdr]
    frac_repertoire = sum(freqs)
    print 'Old frac repertoire:',frac_repertoire
    '''

    pattern = [1 if c in cells_without_scores[:total_matches_at_fdr] else 0 \
               for i,c in enumerate(cells_label)]

    # compare actual frequencies to those that are predicted
    positive_matched_freqs = []
    negative_matched_freqs = []
    non_matched_freqs = []
    positive_confidence = []

    pos_freq_dict = dict([(c[0],(c[1],c[2]))
        for c in cells_with_scores[:total_matches_at_fdr]])
    neg_freq_dict = dict([(c[0],(c[1],c[2]))
        for c in cells_with_scores[total_matches_at_fdr:]])
    
    # assign positive and negative matches
    for cell,true_freq in zip(cells_label,cells_freqs):
        try:
            positive_matched_freqs.append((true_freq,pos_freq_dict[cell][1]['ij']))
            positive_confidence.append((pos_freq_dict[cell][0]))
        except KeyError:
            try:
                negative_matched_freqs.append((true_freq,neg_freq_dict[cell][1]['ij']))
            except KeyError:
                non_matched_freqs.append((true_freq,true_freq))
                #print '{} did not show up!'.format(cell)

    results = {
              'positive_matched_freqs':positive_matched_freqs,
              'negative_matched_freqs':negative_matched_freqs,
              'non_matched_freqs':non_matched_freqs,
              'positive_confidence':positive_confidence,
              'pattern':pattern,
              'frac_repertoire':frac_repertoire,
              'positives':sum(pattern),
              'negatives':len(pattern) - sum(pattern),
              'total':len(pattern),
              'freqs':cells_freqs,
              'xy':[x1,y1],
              'raw_results': cells_with_scores
              }

    if options['save_to_excel']:
        write_results_to_xslx(results)

    if not options['silent']:

        print 'Positives:',results['positives']
        print 'Negatives:',results['negatives']
        print 'Fraction of repertoire: {}%'.format(round(100*results['frac_repertoire'],1))
        print 'Total clones:',results['total']

    return results

#------------------------------------------------------------------------------# 

# OUTDATED, should be unused...
def visualize_results(results,data,**kwargs):

    """
    Processes results dictionary using data object as reference
    """

    # default options parameters
    options = default_options.copy()

    # update options
    options.update(kwargs)

    # heavy lifting on the data
    cresults = analyze_results(results,data,**options)

    ### AUROC FIGURE ###
    plots.plot_auroc(cresults,**kwargs)    

    ### FREQUENCY ESTIMATION FIGURE ###
    plots.plot_frequency_estimation(cresults,**kwargs)    

    ### REPERTOIRE DISPLAY FIGURE ###
    plots.plot_repertoire(cresults,**kwargs)    

    return cresults

#------------------------------------------------------------------------------# 

def is_reference_valid(options):

    reference = options['reference'] # transfer to local namespace

    passing = True

    # if this is not a simulated dataset with a hard set of 
    if reference == None:
        print 'No cell or subject reference provided, cannot produce meaningful results!'
        return False  

    # check to make sure the right dict keys exist
    for s in 'AB':

        if not s in reference:
            print '{} not in reference dict!'.format(s)
            passing = False
            continue

        for c in 'XY':
            if not c in reference[s]:
                print '{} not in reference[{}] subdict!'.format(c,s)
                passing = False

    return passing
    
    check_subject_reference(options)


def compare_results(cresults,**kwargs):

    plots.plot_comparison(cresults,**kwargs) 

